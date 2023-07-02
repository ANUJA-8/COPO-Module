import os
from flask import Flask, render_template, request, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

@app.route('/') 
def index():
    return render_template('index.html')

@app.route('/update_sheet', methods=['POST'])
def update_sheet():
    if request.method == 'POST':
        # Get the uploaded file
        file = request.files['file']

        # Load the workbook
        workbook = load_workbook(file)
        sheet = workbook.active
        sheet1 = workbook.active
        sheet= workbook.get_sheet_by_name('Sheet1')
        sheet1 = workbook.get_sheet_by_name('Sheet2')
        
    # Update the sheet with predefined formulas
    # Sheet1 CO-Attainment-1
    sheet['D80'] = 'target >50'
    sheet['K80'] = 'CO1'
    sheet['M80'] = 'CO2'
    sheet['O80'] = 'CO3'
    sheet['Q80'] = 'CO4'
    sheet['S80'] = 'CO5'
    sheet['U80'] = 'CO6'

    sheet['F86'] = 'CO1'
    sheet['G86'] = 'CO2'
    sheet['H86'] = 'CO3'
    sheet['I86'] = 'CO4'
    sheet['J86'] = 'CO5'
    sheet['K86'] = 'CO6'

    sheet['D86'] = 'Course Outcomes'
    sheet['D87'] = 'Internal Attainment'
    sheet['D88'] = 'External Attainment'
    sheet['D89'] = 'Internal Attainment with 20%'
    sheet['D90'] = 'External Attainment with 80%'
    sheet['D91'] = 'Overall Attainment 80:20'
    sheet['D92'] = 'Overall Attainment 80%'

    # Setting Font    
    font = Font(name='Calibri', size=11, bold=True, italic=False, color='000000')
    for row in sheet['D80:U80']:
        for cell in row:
            cell.font = font

    font = Font(name='Calibri', size=11, bold=True, italic=False, color='000000')
    for row in sheet['F86:K86']:
        for cell in row:
            cell.font = font

    font = Font(name='Calibri', size=11, bold=True, italic=True, color='000000')
    for row in sheet['D86:D92']:
        for cell in row:
            cell.font = font

    sheet["F80"] = '=COUNTIF(F13:F79, ">=50")'
    sheet["K81"] = '=COUNTIF(K13:K79, ">5")'
    sheet["M81"] = '=COUNTIF(M13:M79, ">5")'
    sheet["O81"] = '=COUNTIF(O13:O79, ">5")'
    sheet["Q81"] = '=COUNTIF(Q13:Q79, ">5")'
    sheet["S81"] = '=COUNTIF(S13:S79, ">5")'
    sheet["U81"] = '=COUNTIF(U13:U79, ">5")'

    # Internal Attainment
    sheet['F87'] = "=K81*100/C9"
    sheet['G87'] = "=M81*100/C9"
    sheet['H87'] = "=O81*100/C9"
    sheet['I87'] = "=Q81*100/C9"
    sheet['J87'] = "=S81*100/C9"
    sheet['K87'] = "=U81*100/C9"

    # External Attainment
    sheet['F88'] = "=F80*100/C9"
    sheet['G88'] = "=F80*100/C9"
    sheet['H88'] = "=F80*100/C9"
    sheet['I88'] = "=F80*100/C9"
    sheet['J88'] = "=F80*100/C9"
    sheet['K88'] = "=F80*100/C9"

    # External Attainment with 20%
    sheet['F89'] = "=F87*20/100"
    sheet['G89'] = "=F87*20/100"
    sheet['H89'] = "=F87*20/100"
    sheet['I89'] = "=F87*20/100"
    sheet['J89'] = "=F87*20/100"
    sheet['K89'] = "=F87*20/100"

    # External Attainment with 80%
    sheet['F90'] = "=F88*0.8"
    sheet['G90'] = "=G88*0.8"
    sheet['H90'] = "=H88*0.8"
    sheet['I90'] = "=I88*0.8"
    sheet['J90'] = "=J88*0.8"
    sheet['K90'] = "=K88*0.8"

    # Overall Attainment 80:20
    sheet['F91'] = "=F89+F90"
    sheet['G91'] = "=G89+G90"
    sheet['H91'] = "=H89+H90"
    sheet['I91'] = "=I89+I90"
    sheet['J91'] = "=J89+J90"
    sheet['K91'] = "=K89+K90"

    # Overall Attainment 80%
    sheet['F92'] = "=F91*0.8"
    sheet['G92'] = "=G91*0.8"
    sheet['H92'] = "=H91*0.8"
    sheet['I92'] = "=I91*0.8"
    sheet['J92'] = "=J91*0.8"
    sheet['K92'] = "=K91*0.8"

    # Setting Dimensions
    sheet.column_dimensions['D'].width = 20

    # Setting allignement
    for row in sheet['D80:U81']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adding Border
    def set_border(sheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    set_border(sheet, 'D80:U81') 

    def set_border(sheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    set_border(sheet, 'D86:K92')

    # Adding Colour
    _color1 = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    for row in sheet['D86:K92']:
        for cell in row:
            cell.fill = _color1

        _color2 = PatternFill(start_color="808000", end_color="808000", fill_type="solid")
    for row in sheet['D80:U81']:
        for cell in row:
            cell.fill = _color2

    # Sheet-2 Indirect Shift-1
    sheet1['A50'] = 'CO1'
    sheet1['A51'] = 'CO2'
    sheet1['A52'] = 'CO3'
    sheet1['A53'] = 'CO4'
    sheet1['A54'] = 'CO5'
    sheet1['A55'] = 'CO6'
    sheet1['A56'] = 'ALL'

    sheet1['B48'] = 5
    sheet1['C48'] = 4
    sheet1['D48'] = 3
    sheet1['E48'] = 2
    sheet1['F48'] = 1

    sheet1['B49'] = 'Excellent'
    sheet1['C49'] = 'Good'
    sheet1['D49'] = 'Average'
    sheet1['E49'] = 'Fair'
    sheet1['F49'] = 'Poor'
    sheet1['G49'] = 'Total Score'
    sheet1['H49'] = '% Score'
    sheet1['I49'] = 'Avg %'
    sheet1['J49'] = '20% Score'

    # Setting Allignment
    for row in sheet1['A48:J56']:
     for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adding color
     _color3 = PatternFill(start_color="993300", end_color="993300", fill_type="solid")
     for row in sheet1['A50:A56']:
        for cell in row:
         cell.fill = _color3

    _color4 = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    for row in sheet1['B48:F48']:
      for cell in row:
        cell.fill = _color4        

    # Excellent
    sheet1["B50"] = '=COUNTIF(C9:C46, ">=5")'
    sheet1["B51"] = '=COUNTIF(D9:D46, ">=5")'
    sheet1["B52"] = '=COUNTIF(E9:E46, ">=5")'
    sheet1["B53"] = '=COUNTIF(F9:F46, ">=5")'
    sheet1["B54"] = '=COUNTIF(G9:G46, ">=5")'
    sheet1["B55"] = '=COUNTIF(H9:H46, ">=5")'
    sheet1["B56"] = '=COUNTIF(I9:I46, ">=5")'

    # Good
    sheet1["C50"] = '=COUNTIF(C9:C46, "=4")'
    sheet1["C51"] = '=COUNTIF(D9:D46, "=4")'
    sheet1["C52"] = '=COUNTIF(E9:E46, "=4")'
    sheet1["C53"] = '=COUNTIF(F9:F46, "=4")'
    sheet1["C54"] = '=COUNTIF(G9:G46, "=4")'
    sheet1["C55"] = '=COUNTIF(H9:H46, "=4")'
    sheet1["C56"] = '=COUNTIF(I9:I46, "=4")'

    #Average
    sheet1["D50"] = '=COUNTIF(C9:C46, "=3")'
    sheet1["D51"] = '=COUNTIF(D9:D46, "=3")'
    sheet1["D52"] = '=COUNTIF(E9:E46, "=3")'
    sheet1["D53"] = '=COUNTIF(F9:F46, "=3")'
    sheet1["D54"] = '=COUNTIF(G9:G46, "=3")'
    sheet1["D55"] = '=COUNTIF(H9:H46, "=3")'
    sheet1["D56"] = '=COUNTIF(I9:I46, "=3")'

    # Fair
    sheet1["E50"] = '=COUNTIF(C9:C46, "=2")'
    sheet1["E51"] = '=COUNTIF(D9:D46, "=2")'
    sheet1["E52"] = '=COUNTIF(E9:E46, "=2")'
    sheet1["E53"] = '=COUNTIF(F9:F46, "=2")'
    sheet1["E54"] = '=COUNTIF(G9:G46, "=2")'
    sheet1["E55"] = '=COUNTIF(H9:H46, "=2")'
    sheet1["E56"] = '=COUNTIF(I9:I46, "=2")'

    # Poor
    sheet1["F50"] = '=COUNTIF(C9:C46, "=1")'
    sheet1["F51"] = '=COUNTIF(D9:D46, "=1")'
    sheet1["F52"] = '=COUNTIF(E9:E46, "=1")'
    sheet1["F53"] = '=COUNTIF(F9:F46, "=1")'
    sheet1["F54"] = '=COUNTIF(G9:G46, "=1")'
    sheet1["F55"] = '=COUNTIF(H9:H46, "=1")'
    sheet1["F56"] = '=COUNTIF(I9:I46, "=1")'

    # Total Score
    sheet1['G50'] = "=B50*5+C50*4+D50*3+E50*2+F50*1"
    sheet1['G51'] = "=B51*5+C51*4+D51*3+E51*2+F51*1"
    sheet1['G52'] = "=B52*5+C52*4+D52*3+E52*2+F52*1"
    sheet1['G53'] = "=B53*5+C53*4+D53*3+E53*2+F53*1"
    sheet1['G54'] = "=B54*5+C54*4+D54*3+E54*2+F54*1"
    sheet1['G55'] = "=B55*5+C55*4+D55*3+E55*2+F55*1"
    sheet1['G56'] = "=B56*5+C56*4+D56*3+E56*2+F56*1"

    # %Score
    sheet1['H50'] = "=G50/(SUM(B50:F50)*.05)"
    sheet1['H51'] = "=G51/(SUM(B51:F51)*.05)"
    sheet1['H52'] = "=G52/(SUM(B52:F52)*.05)"
    sheet1['H53'] = "=G53/(SUM(B53:F53)*.05)"
    sheet1['H54'] = "=G54/(SUM(B54:F54)*.05)"
    sheet1['H55'] = "=G55/(SUM(B55:F55)*.05)"
    sheet1['H56'] = "=G56/(SUM(B56:F56)*.05)"

    # AVG%
    sheet1['I50'] = "=AVERAGE(H50,H56)"
    sheet1['I51'] = "=AVERAGE(H51,H56)"
    sheet1['I52'] = "=AVERAGE(H52,H56)"
    sheet1['I53'] = "=AVERAGE(H53,H56)"
    sheet1['I54'] = "=AVERAGE(H54,H56)"
    sheet1['I55'] = "=AVERAGE(H55,H56)"
    sheet1['I56'] = "=AVERAGE(H56,H56)"

    # 20% Score
    sheet1['J50'] = "=I50*20%"
    sheet1['J51'] = "=I51*20%"
    sheet1['J52'] = "=I52*20%"
    sheet1['J53'] = "=I53*20%"
    sheet1['J54'] = "=I54*20%"
    sheet1['J55'] = "=I55*20%"
    sheet1['J56'] = "=I56*20%"

    # Adding border
    def set_border(sheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in sheet1[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    set_border(sheet1, 'A48:J56')

    # Setting font
    font = Font(name='Calibri', size=11, bold=True, italic=True, color='000000')
    for row in sheet1['B49:J49']:
        for cell in row:
            cell.font = font

    font = Font(name='Calibri', size=11, bold=True, italic=True, color='000000')
    for row in sheet1['A50:A56']:
        for cell in row:
            cell.font = font

        # Save the updated workbook
    filename = 'updated_sheet.xlsx'
    workbook.save(filename)

        # Send the updated sheet as a downloadable file
    return send_from_directory(directory=os.getcwd(), path=filename, as_attachment=True)

    # Return an empty response if the method is not POST
    return ''

if __name__ == '__main__':
    app.run(port=5500)
